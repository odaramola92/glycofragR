"""
High-level unified API for glycan-only analysis and structure prediction.

Single-line initialization that handles everything internally:
- Glycan creation
- Fragment generation
- Theoretical table generation (raw + clean)
- Summary sheet creation
- Optional visualization

Example:
    >>> from glycofrag import GlycanAnalysis
    >>> 
    >>> # Everything in one call
    >>> analysis = GlycanAnalysis(
    ...     glycan_code='4501',
    ...     glycan_type='N',
    ...     modification_type='permethylated_reduced',
    ...     fragment_types=['BY', 'CZ'],
    ...     visualize_structure='all'  # or 'best', [1,2,3], None
    ... )
    >>> 
    >>> # Access generated data
    >>> print(analysis.theoretical_df)  # raw theoretical fragments
    >>> print(analysis.clean_theoretical_df)  # deduplicated
    >>> print(analysis.summary_df)  # mass information
    >>> 
    >>> # For MS/MS matching (released glycan analysis)
    >>> matched = analysis.match_fragments(spectrum, ppm_tolerance=20)
    >>> prediction = analysis.predict_structures(matched)
    >>> analysis.visualize_best_structure('structure.png')
    >>> 
    >>> # Export to Excel
    >>> analysis.export_to_excel('glycan_analysis.xlsx', matched)
"""

import os
from typing import Optional, Dict, List, Tuple, Union
import numpy as np
import pandas as pd
from glycofrag.glycan.facade import Glycan
from glycofrag.io.tables import (
    build_clean_theoretical,
    score_structures_by_uniqueness,
    create_structure_prediction_sheet,
    get_best_structure,
    generate_fragment_table,
    match_fragments_from_table,
)
from glycofrag.io.visualizer import GlycanVisualizer
from glycofrag.core.mass_calculator import GlycanMassCalculator


class GlycanAnalysis:
    """
    All-in-one interface for complete glycan MS/MS analysis pipeline.
    
    Initialization handles ALL setup:
    - Creates Glycan object
    - Predicts structures
    - Generates fragment ions
    - Builds theoretical fragment tables (raw + deduplicated)
    - Creates summary sheet with masses/modifications
    - Optionally visualizes structures
    
    Then provides simple methods for:
    - Fragment matching (experimental)
    - Structure prediction/ranking
    - Visualization and export
    
    Example:
        >>> analysis = GlycanAnalysis(
        ...     glycan_code='4501',
        ...     glycan_type='N',
        ...     modification_type='permethylated_reduced',
        ...     fragment_types=['BY'],
        ...     visualize_structure='best'
        ... )
        >>>
        >>> # Access all generated data
        >>> print(analysis.theoretical_df)  # all fragments
        >>> print(analysis.clean_theoretical_df)  # deduplicated
        >>> print(analysis.mass)  # glycan neutral mass
        >>>
        >>> # Match experimental peaks
        >>> matched_df = analysis.match_fragments(spectrum, ppm_tolerance=20)
        >>>
        >>> # Rank structures
        >>> prediction = analysis.predict_structures(matched_df)
        >>> analysis.visualize_best_structure('best_structure.png')
        >>> analysis.export_to_excel('analysis.xlsx', matched_df)
    """
    
    def __init__(
        self,
        glycan_code: str,
        glycan_type: str = 'N',
        modification_type: Union[str, int] = 'free',
        max_structures: int = 100,
        isomer_sensitive: bool = False,
        fragment_types: Optional[List[str]] = None,
        charges: Optional[List[int]] = None,
        visualize_structure: Optional[Union[str, int, List[int]]] = None,
        preferred_core: Optional[int] = None,
        custom_reducing_end_mass: Optional[float] = None,
        custom_mono_masses: Optional[Dict[str, float]] = None,
    ):
        """
        Initialize and run complete glycan analysis pipeline.
        
        Args:
            glycan_code: Glycan composition code (e.g., '4501')
            glycan_type: 'N' or 'O' for N/O-glycan (default 'N')
            modification_type: Reducing end modification type
                String options (recommended):
                  - 'free': Free reducing end (default)
                  - 'reduced': Reduced/alditol end
                  - 'permethylated_free' or 'pm_free': Permethylated free
                  - 'permethylated_reduced' or 'pm_reduced': Permethylated reduced
                  - '2ab': 2-AB labeled
                  - '2ab_permethylated': 2-AB + permethylated
                  - 'custom': Custom (requires custom_reducing_end_mass)
                Integer options: 0-5, 7 (see Glycan class docs)
            max_structures: Max structures to predict (default 100)
            isomer_sensitive: If True, treats mirror images as distinct
            fragment_types: Fragment series to generate
                Options: ['BY'], ['CZ'], or ['BY', 'CZ']
                Default: ['BY']
            charges: Charge states for m/z calculation (default [1, 2, 3])
            visualize_structure: Which structures to visualize
                - 'best': Top-ranked structure (after matching)
                - 'all': All predicted structures
                - int (1, 2, 3...): Specific structure
                - list [1, 2, 3]: Multiple structures
                - None: No visualization
            preferred_core: For O-glycans with ambiguous cores.
                Specify which core to use: 1, 3, 5, 6, 7, or 8.
                If None, defaults to lowest core number.
            custom_reducing_end_mass: Mass to add at reducing end (for custom type)
            custom_mono_masses: Dict of masses to add per monosaccharide (for custom type)
        
        Generates and stores internally:
            - theoretical_df: All theoretical fragments (raw)
            - clean_theoretical_df: Deduplicated by mass+composition
            - summary_df: Masses, modifications, QC info
            - structures: Glycan structure objects
        """
        # Store parameters
        self.glycan_code = glycan_code
        self.glycan_type = glycan_type
        self.modification_type = modification_type
        self.charges = charges if charges is not None else [1, 2, 3]
        
        # DEBUG: Print what we received
        print(f"DEBUG: fragment_types received = {fragment_types!r}")
        print(f"DEBUG: fragment_types type = {type(fragment_types)}")
        if fragment_types:
            print(f"DEBUG: fragment_types is truthy, length = {len(fragment_types)}")
        else:
            print(f"DEBUG: fragment_types is falsy!")
        
        # === STEP 1: Create Glycan (internal) ===
        self.glycan = Glycan(
            glycan_code,
            glycan_type=glycan_type,
            max_structures=max_structures,
            modification_type=modification_type,
            isomer_sensitive=isomer_sensitive,
            preferred_core=preferred_core,
            custom_reducing_end_mass=custom_reducing_end_mass,
            custom_mono_masses=custom_mono_masses,
        )
        
        # === STEP 2: Predict structures ===
        self.structures = self.glycan.predict_structures()
        
        # === STEP 3: Generate fragments (if requested) ===
        self.fragments_by_structure = []
        if fragment_types:
            for structure in self.structures:
                fragments, cleavage_info = self.glycan.generate_fragments(
                    structure=structure,
                    fragment_types=fragment_types,
                    charges=self.charges,
                )
                self.fragments_by_structure.append({
                    'fragments': fragments,
                    'cleavage_info': cleavage_info,
                    'structure': structure
                })
        
        # === STEP 4: Build theoretical tables (internal) ===
        if self.fragments_by_structure:
            # Aggregate all fragments across structures
            all_fragments = self._aggregate_fragments()
            self.theoretical_df = generate_fragment_table(
                fragments=all_fragments,
                glycan_code=glycan_code,
                modification_type=modification_type,
                charge_states=self.charges,
            )
        else:
            self.theoretical_df = pd.DataFrame()
        
        # Deduplicate by mass + composition
        self.clean_theoretical_df = build_clean_theoretical(
            self.theoretical_df, 
            mass_round=6
        ) if not self.theoretical_df.empty else pd.DataFrame()
        
        # === STEP 5: Create summary sheet (internal) ===
        self.mass_calculator = GlycanMassCalculator(modification_type=self.glycan.modification_type)
        glycan_mass = self.mass_calculator.calculate_glycan_mass(glycan_code)
        
        self.summary_df = self._create_glycan_summary(
            glycan_code=glycan_code,
            glycan_type=glycan_type,
            glycan_mass=glycan_mass,
            modification_type=modification_type,
        )
        
        # Cache for analysis results
        self.prediction_df: Optional[pd.DataFrame] = None
        self.scores: Optional[Dict] = None
        
        # === STEP 6: Optional visualization ===
        if visualize_structure and self.structures:
            self._auto_visualize(visualize_structure)
    
    def _aggregate_fragments(self) -> Dict:
        """Aggregate fragments from all structures with structure tracking."""
        aggregated = {}
        
        for idx, frag_data in enumerate(self.fragments_by_structure, start=1):
            fragments = frag_data['fragments']
            
            for frag_type, frag_list in fragments.items():
                if frag_type not in aggregated:
                    aggregated[frag_type] = []
                
                # Add structure number to each fragment
                for frag in frag_list:
                    frag_copy = frag.copy()
                    frag_copy['structure'] = idx
                    aggregated[frag_type].append(frag_copy)
        
        return aggregated
    
    def _create_glycan_summary(
        self, 
        glycan_code: str, 
        glycan_type: str,
        glycan_mass: float,
        modification_type: Union[str, int]
    ) -> pd.DataFrame:
        """Create summary DataFrame with glycan information."""
        
        # Get modification name
        from glycofrag.core.modifications import get_modification_name
        mod_name = get_modification_name(modification_type)
        
        summary_data = {
            'Parameter': [
                'Glycan Code',
                'Glycan Type',
                'Reducing End Modification',
                'Neutral Mass (Da)',
                'Number of Structures Predicted',
            ],
            'Value': [
                glycan_code,
                glycan_type,
                mod_name,
                f"{glycan_mass:.6f}",
                len(self.structures),
            ]
        }
        
        return pd.DataFrame(summary_data)
    
    def _auto_visualize(self, option: Union[str, int, List[int]]):
        """Handle automatic visualization based on option."""
        if option == 'best':
            # Visualize first structure (or best after matching)
            if self.structures:
                GlycanVisualizer.visualize_structure(
                    self.structures,
                    structure_number=1,
                    show=False,
                    output_path=f'glycan_{self.glycan_code}_structure_1.png'
                )
        elif option == 'all':
            # Visualize all structures
            for i in range(1, len(self.structures) + 1):
                GlycanVisualizer.visualize_structure(
                    self.structures,
                    structure_number=i,
                    show=False,
                    output_path=f'glycan_{self.glycan_code}_structure_{i}.png'
                )
        elif isinstance(option, int):
            # Visualize specific structure
            if 1 <= option <= len(self.structures):
                GlycanVisualizer.visualize_structure(
                    self.structures,
                    structure_number=option,
                    show=False,
                    output_path=f'glycan_{self.glycan_code}_structure_{option}.png'
                )
        elif isinstance(option, list):
            # Visualize multiple structures
            for i in option:
                if 1 <= i <= len(self.structures):
                    GlycanVisualizer.visualize_structure(
                        self.structures,
                        structure_number=i,
                        show=False,
                        output_path=f'glycan_{self.glycan_code}_structure_{i}.png'
                    )
    
    @property
    def mass(self) -> float:
        """Get neutral mass of the glycan."""
        return self.mass_calculator.calculate_glycan_mass(self.glycan_code)
    
    def match_fragments(
        self,
        experimental_spectrum: Union[pd.DataFrame, List[Tuple[float, float]], np.ndarray],
        ppm_tolerance: float = 20.0,
        intensity_threshold: float = 0.0
    ) -> pd.DataFrame:
        """
        Match experimental spectrum against theoretical fragments.
        
        Args:
            experimental_spectrum: Experimental m/z and intensity data
                Can be: DataFrame with 'mz' and 'intensity' columns,
                       List of (mz, intensity) tuples,
                       or numpy array of shape (N, 2)
            ppm_tolerance: Mass tolerance in ppm (default 20)
            intensity_threshold: Minimum relative intensity (0-100, default 0)
        
        Returns:
            DataFrame with matched fragments including ppm error
        """
        if self.theoretical_df.empty:
            raise ValueError("No theoretical fragments available. Generate fragments first.")
        
        matched_df = match_fragments_from_table(
            theoretical_df=self.theoretical_df,
            experimental_spectrum=experimental_spectrum,
            ppm_tolerance=ppm_tolerance,
            intensity_threshold=intensity_threshold
        )
        
        return matched_df
    
    def predict_structures(self, matched_df: pd.DataFrame) -> pd.DataFrame:
        """
        Rank structures based on matched fragments.
        
        Args:
            matched_df: DataFrame from match_fragments()
        
        Returns:
            DataFrame with structure rankings and scores
        """
        if matched_df.empty:
            raise ValueError("No matched fragments provided")
        
        # Score structures
        self.scores = score_structures_by_uniqueness(
            matched_df=matched_df,
            num_structures=len(self.structures)
        )
        
        # Create prediction DataFrame
        self.prediction_df = create_structure_prediction_sheet(
            scores=self.scores,
            structures=self.structures,
            glycan=self.glycan
        )
        
        return self.prediction_df
    
    def visualize_best_structure(
        self,
        output_path: Optional[str] = None,
        show: bool = False,
        show_node_numbers: bool = True,
        **kwargs
    ):
        """
        Visualize the best-ranked structure (after predict_structures).
        
        Args:
            output_path: Path to save image (default: glycan_{code}_best.png)
            show: Display plot interactively
            show_node_numbers: Show node numbers on structure
            **kwargs: Advanced visualization options passed to GlycanVisualizer:
                - vertical_gap (float | dict): Vertical spacing between levels
                - horizontal_spacing (float | dict): Horizontal spacing between siblings
                - node_size (float | dict): Monosaccharide marker sizes
                - title (str): Figure title
                - figsize (tuple): Figure size in inches
        """
        if self.prediction_df is None:
            # No ranking yet, use first structure
            best_idx = 1
        else:
            best_idx = get_best_structure(self.prediction_df)
        
        if output_path is None:
            output_path = f'glycan_{self.glycan_code}_best_structure.png'
        
        GlycanVisualizer.visualize_structure(
            self.structures,
            structure_number=best_idx,
            show=show,
            output_path=output_path,
            show_node_numbers=show_node_numbers,
            **kwargs
        )
    
    def visualize_structure(
        self,
        structure_number: int,
        output_path: Optional[str] = None,
        show: bool = False,
        show_node_numbers: bool = True,
        **kwargs
    ):
        """
        Visualize a specific structure by number.
        
        Args:
            structure_number: Structure index (1-based)
            output_path: Path to save image
            show: Display plot interactively
            show_node_numbers: Show node numbers on structure
            **kwargs: Advanced visualization options passed to GlycanVisualizer:
                - vertical_gap (float | dict): Vertical spacing between levels
                - horizontal_spacing (float | dict): Horizontal spacing between siblings
                - node_size (float | dict): Monosaccharide marker sizes
                - title (str): Figure title
                - figsize (tuple): Figure size in inches
        
        Example:
            >>> analysis.visualize_structure(
            ...     1,
            ...     vertical_gap={3: 2.0, 4: 2.5},
            ...     horizontal_spacing=2.5,
            ...     node_size={'HexNAc': 350}
            ... )
        """
        if structure_number < 1 or structure_number > len(self.structures):
            raise ValueError(f"Structure number must be 1-{len(self.structures)}")
        
        if output_path is None:
            output_path = f'glycan_{self.glycan_code}_structure_{structure_number}.png'
        
        GlycanVisualizer.visualize_structure(
            self.structures,
            structure_number=structure_number,
            show=show,
            output_path=output_path,
            show_node_numbers=show_node_numbers,
            **kwargs
        )
    
    def export_to_excel(
        self,
        output_file: str,
        matched_df: Optional[pd.DataFrame] = None,
        include_prediction: bool = True
    ):
        """
        Export all analysis results to Excel workbook.
        
        Args:
            output_file: Path to save Excel file
            matched_df: Optional matched fragments DataFrame
            include_prediction: Include structure prediction sheet
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # === Summary Sheet ===
        ws_summary = wb.create_sheet("Summary", 0)
        ws_summary['A1'] = "Glycan Analysis Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)
        
        for idx, row in self.summary_df.iterrows():
            ws_summary[f'A{idx+3}'] = row['Parameter']
            ws_summary[f'B{idx+3}'] = row['Value']
            ws_summary[f'A{idx+3}'].font = Font(bold=True)
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        # === Theoretical Fragments Sheet ===
        if not self.theoretical_df.empty:
            ws_theo = wb.create_sheet("Theoretical Fragments")
            
            # Write headers
            for col_idx, col_name in enumerate(self.theoretical_df.columns, 1):
                cell = ws_theo.cell(1, col_idx, col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            # Write data
            for row_idx, row in self.theoretical_df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws_theo.cell(row_idx + 2, col_idx, value)
            
            # Adjust column widths
            for col_idx in range(1, len(self.theoretical_df.columns) + 1):
                ws_theo.column_dimensions[chr(64 + col_idx)].width = 15
        
        # === Clean Theoretical Sheet ===
        if not self.clean_theoretical_df.empty:
            ws_clean = wb.create_sheet("Clean Theoretical")
            
            for col_idx, col_name in enumerate(self.clean_theoretical_df.columns, 1):
                cell = ws_clean.cell(1, col_idx, col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            for row_idx, row in self.clean_theoretical_df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws_clean.cell(row_idx + 2, col_idx, value)
            
            for col_idx in range(1, len(self.clean_theoretical_df.columns) + 1):
                ws_clean.column_dimensions[chr(64 + col_idx)].width = 15
        
        # === Matched Fragments Sheet ===
        if matched_df is not None and not matched_df.empty:
            ws_matched = wb.create_sheet("Matched Fragments")
            
            for col_idx, col_name in enumerate(matched_df.columns, 1):
                cell = ws_matched.cell(1, col_idx, col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            for row_idx, row in matched_df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws_matched.cell(row_idx + 2, col_idx, value)
            
            for col_idx in range(1, len(matched_df.columns) + 1):
                ws_matched.column_dimensions[chr(64 + col_idx)].width = 15
        
        # === Structure Prediction Sheet ===
        if include_prediction and self.prediction_df is not None and not self.prediction_df.empty:
            ws_pred = wb.create_sheet("Structure Prediction")
            
            for col_idx, col_name in enumerate(self.prediction_df.columns, 1):
                cell = ws_pred.cell(1, col_idx, col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            
            for row_idx, row in self.prediction_df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws_pred.cell(row_idx + 2, col_idx, value)
            
            for col_idx in range(1, len(self.prediction_df.columns) + 1):
                ws_pred.column_dimensions[chr(64 + col_idx)].width = 20
        
        # Save workbook
        wb.save(output_file)
        print(f"[OK] Analysis exported to: {output_file}")
    
    # ========== R-COMPATIBLE CONVERSION METHODS ==========
    # These methods convert pandas DataFrames to R-friendly formats
    # (list of dictionaries) for seamless interchange with R via reticulate
    
    def get_theoretical_df_as_list(self) -> List[Dict]:
        """
        Convert theoretical_df to R-compatible list of dictionaries.
        
        Each row becomes a dictionary, making it easy for R to work with.
        Handles NaN/None values appropriately.
        
        Returns:
            List of dictionaries, one per row in theoretical_df.
            Empty list if no theoretical fragments.
        """
        if self.theoretical_df.empty:
            return []
        return self.theoretical_df.fillna("").to_dict('records')
    
    def get_clean_df_as_list(self) -> List[Dict]:
        """
        Convert clean_theoretical_df to R-compatible list of dictionaries.
        
        Each row becomes a dictionary, making it easy for R to work with.
        Handles NaN/None values appropriately.
        
        Returns:
            List of dictionaries, one per row in clean_theoretical_df.
            Empty list if no clean fragments.
        """
        if self.clean_theoretical_df.empty:
            return []
        return self.clean_theoretical_df.fillna("").to_dict('records')
    
    def get_summary_df_as_list(self) -> List[Dict]:
        """
        Convert summary_df to R-compatible list of dictionaries.
        
        Each row becomes a dictionary, making it easy for R to work with.
        Handles NaN/None values appropriately.
        
        Returns:
            List of dictionaries, one per row in summary_df.
            Empty list if no summary data.
        """
        if self.summary_df.empty:
            return []
        return self.summary_df.fillna("").to_dict('records')
    
    def get_structures_as_list(self) -> List[Dict]:
        """
        Convert predicted structures to R-compatible format.
        
        Returns:
            List of dictionaries representing each predicted structure.
            Each dict contains structure information and string representation.
        """
        structures_list = []
        for idx, struct in enumerate(self.structures, 1):
            structures_list.append({
                'structure_id': idx,
                'structure_str': str(struct),
                'structure_repr': repr(struct)
            })
        return structures_list
    
    def __repr__(self) -> str:
        """String representation of the analysis."""
        return (
            f"GlycanAnalysis(\n"
            f"  glycan_code='{self.glycan_code}',\n"
            f"  glycan_type='{self.glycan_type}',\n"
            f"  modification_type={self.modification_type},\n"
            f"  structures_predicted={len(self.structures)},\n"
            f"  theoretical_fragments={len(self.theoretical_df)},\n"
            f"  unique_fragments={len(self.clean_theoretical_df)}\n"
            f")"
        )
